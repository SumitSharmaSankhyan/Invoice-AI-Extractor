from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
import os

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

last_result = []


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/extract", methods=["POST"])
def extract():

    global last_result

    files = request.files.getlist("files")

    results = []

    for file in files:

        filename = file.filename

        save_path = os.path.join(UPLOAD_FOLDER, filename)

        file.save(save_path)

        # --------------------------
        # TEMPORARY SAMPLE DATA
        # OCR will replace this later
        # --------------------------

        results.append({
            "po": "PO123456",
            "invoice": "INV001",
            "date": "25-06-2026",
            "base": "10000",
            "igst": "1800",
            "cgst": "0",
            "sgst": "0",
            "other": "0",
            "total": "11800"
        })

    last_result = results

    return jsonify(results)


@app.route("/download")
def download():

    wb = Workbook()

    ws = wb.active

    ws.title = "Data Invoice"

    # Column Headers
    ws["A1"] = "PO Code"

    ws["W1"] = "Invoice No"

    ws["X1"] = "Invoice Date"

    ws["Y1"] = "Base Amount"

    ws["Z1"] = "IGST"

    ws["AA1"] = "CGST"

    ws["AB1"] = "SGST"

    ws["AC1"] = "Other Charges"

    ws["AD1"] = "Total Amount"

    row = 2

    for item in last_result:

        ws[f"A{row}"] = item["po"]

        ws[f"W{row}"] = item["invoice"]

        ws[f"X{row}"] = item["date"]

        ws[f"Y{row}"] = item["base"]

        ws[f"Z{row}"] = item["igst"]

        ws[f"AA{row}"] = item["cgst"]

        ws[f"AB{row}"] = item["sgst"]

        ws[f"AC{row}"] = item["other"]

        ws[f"AD{row}"] = item["total"]

        row += 1

    output_path = os.path.join(
        OUTPUT_FOLDER,
        "Invoice_Output.xlsx"
    )

    wb.save(output_path)

    return send_file(
        output_path,
        as_attachment=True
    )


if __name__ == "__main__":
    app.run(debug=True)
